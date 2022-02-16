import inspect
import softest
import logging
import csv
from openpyxl import Workbook, load_workbook

class Utils(softest.TestCase):
    def assert_list_item_text(self,list,value):
        for an_item in list:
            self.soft_assert(self.assertEqual,an_item, value)
            if an_item==value:
                print("assert pass")
            else:
                print("assert fail")
        self.assert_all()

    def custom_logger(loglevel=logging.DEBUG):
        # create logger
        logger_name=inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)

        # create console handler and set level to debug
        # ch = logging.StreamHandler()
        # ch.setLevel(loglevel)

        fh=logging.FileHandler("automation.log", mode='w')

        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',\
                                      datefmt='%m/%d/%Y %I:%M:%S %p')

        # add formatter to ch
        fh.setFormatter(formatter)

        # add ch to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist=[]
        wb= load_workbook(filename=file_name)
        sh=wb[sheet]
        row_count=sh.max_row
        col_count=sh.max_column

        for i in range(2,row_count+1):
            row=[]
            for j in range (1, col_count+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist=[]
        csvdata= open(filename,"r")
        reader= csv.reader(csvdata)
        next (reader)
        for row in reader:
            datalist.append(row)
        return datalist
















